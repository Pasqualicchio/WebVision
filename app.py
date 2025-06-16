from flask import Flask, request, jsonify
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Inizializza l'app Flask
app = Flask(__name__)

# Configurazione
UPLOAD_FOLDER = 'static/uploads/'
EXCEL_FILE = 'static/registro_foto.xlsx'
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'mp4'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Verifica che il file abbia un'estensione consentita
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 🏠 Homepage di test
@app.route('/')
def home():
    return "✅ WebVision API è attiva e funzionante!"

# 🔐 Login di base
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    if data.get('username') == 'admin' and data.get('password') == 'admin123':
        return jsonify({"message": "Login riuscito!"}), 200
    return jsonify({"message": "Credenziali non valide"}), 401

# 📤 Upload dei file + aggiornamento Excel
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"message": "Nessun file inviato"}), 400

    file = request.files['file']
    
    if file.filename == '':
        return jsonify({"message": "Nessun file selezionato"}), 400

    if not allowed_file(file.filename):
        return jsonify({"message": "Tipo di file non consentito"}), 400

    # Crea la cartella di upload se non esiste
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    # Salva il file
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # Aggiorna il file Excel
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Data', 'Utente', 'Commessa', 'Tipo file', 'Percorso', 'Commenti', 'GPS'])

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([current_time, 'admin', 'Commessa001', file.filename, file_path, 'Nessun commento', 'GPS_info'])
        wb.save(EXCEL_FILE)

    except Exception as e:
        return jsonify({"message": f"Errore nell'aggiornamento del file Excel: {str(e)}"}), 500

    return jsonify({"message": f"✅ File '{file.filename}' caricato con successo!"}), 200

# 🚀 Avvio server compatibile con Render
if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))  # Render imposta automaticamente la porta
    app.run(
        host='0.0.0.0',  # Accetta connessioni da qualunque IP
        port=port,
        debug=True,
        use_reloader=False
    )